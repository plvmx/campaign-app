#!/usr/bin/env python3
"""
Campaign Report project — see docs/campaign-report/BRIEF.md.

Converts the "campaign report" sheet of an AFJ Tracking Export (Jordan's
Google Sheet export, .xlsx) into a JSON array that
scripts/import_campaign_reports.ts can read. Used for both the initial
historical load and any later incremental catch-up export.

Why Python instead of an npm xlsx-parsing package: as of this writing, both
`xlsx` (SheetJS) and `exceljs` on the npm registry carry unpatched high/
moderate-severity advisories (prototype pollution / ReDoS) — see the
"docs in sync" note in this project's PR. openpyxl has none, and it's what
this file was already extracted with during the investigation this project
grew out of, so it's the extraction step of record rather than a two-line
convenience import.

Each output row is an object keyed by the target field name, with every
value tagged by type so the TypeScript side (which never imports openpyxl or
any xlsx library) can reconstruct it exactly without guessing:
  {"t": "date",   "v": "2021-09-02T18:17:36"}   -- native Excel date/time
  {"t": "number", "v": 20240914}                -- a plain number
  {"t": "string", "v": "18th September 2021"}   -- text
  {"t": "null",   "v": null}                    -- empty cell

Usage:
  python3 scripts/campaign_reports_xlsx_to_json.py <input.xlsx> <output.json>

Requirements: openpyxl (pip install openpyxl if the import below fails).
"""
import datetime
import json
import sys

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "campaign report"


def tag(value):
    if value is None:
        return {"t": "null", "v": None}
    if isinstance(value, (datetime.datetime, datetime.date)):
        return {"t": "date", "v": value.isoformat()}
    if isinstance(value, datetime.time):
        # A time-of-day with no date part landed in a date column by mistake
        # (seen once in the real sheet) — unrecoverable as a date, pass
        # through as text so the parser flags it for review rather than
        # silently dropping it.
        return {"t": "string", "v": f"[time-only: {value}]"}
    if isinstance(value, (int, float)):
        return {"t": "number", "v": value}
    return {"t": "string", "v": str(value)}


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(1)
    input_path, output_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"No '{SHEET_NAME}' sheet found. Sheets present: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET_NAME]

    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        rows_out.append({
            "submitted_at": tag(row[0]),
            "location": tag(row[1]),
            "leader": tag(row[2]),
            "campaign_date": tag(row[3]),
            "partial_presentations": tag(row[4]),
            "full_presentations": tag(row[5]),
            "sinners_prayer": tag(row[6]),
            "information_requests": tag(row[7]),
        })

    with open(output_path, "w") as f:
        json.dump(rows_out, f)

    print(f"Wrote {len(rows_out)} rows to {output_path}")


if __name__ == "__main__":
    main()
